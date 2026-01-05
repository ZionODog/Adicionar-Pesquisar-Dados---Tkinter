import tkinter as tk
from tkinter import messagebox
import openpyxl
import os

def abrir_janela_busca(master_window=None):
    # Se vier do menu principal usa Toplevel, se rodar sozinho usa Tk
    if master_window:
        janela = tk.Toplevel(master_window)
    else:
        janela = tk.Tk()

    janela.title("Pesquisar no Relatório")
    janela.geometry("800x500")
    
    # --- CAMINHO DO ARQUIVO ---
    # 1. Pega a pasta onde o script está (pasta 'modulos')
    dir_modulos = os.path.dirname(os.path.abspath(__file__))
    
    # 2. Sobe um nível para a raiz do projeto ('EXCEL-TOOLS-SUITE')
    dir_raiz = os.path.dirname(dir_modulos)
    
    # 3. Monta o caminho descendo para a pasta 'data'
    caminho_excel = os.path.join(dir_raiz, 'data', 'Relatorio.xlsx')
    
    # 4. Monta o caminho descendo para a pasta 'theme' (para o cadastro.py e main.py)
    caminho_tema_dark = os.path.join(dir_raiz, 'theme', 'forest-dark.tcl')
    caminho_tema_light = os.path.join(dir_raiz, 'theme', 'forest-light.tcl')

    def submit():
        termo_busca = entry_colaborador.get()
        
        # Habilita edição para poder escrever os resultados e depois trava de novo
        ativar_campos()
        limpar_resultados()
        
        if not termo_busca:
            messagebox.showwarning("Aviso", "Digite o nome do Colaborador para buscar.")
            desativar_campos()
            return

        if not os.path.exists(caminho_excel):
            messagebox.showerror("Erro", f"Arquivo não encontrado!")
            desativar_campos()
            return

        try:
            # Carrega a planilha
            workbook = openpyxl.load_workbook(caminho_excel)
            # Tenta pegar a aba 'Ativos', se não der erro, tenta a ativa
            if "Ativos" in workbook.sheetnames:
                sheet = workbook["Ativos"]
            else:
                sheet = workbook.active
            
            encontrado = False

            # Itera sobre as linhas (pulando o cabeçalho)
            # Colunas: 0=Colaborador, 1=Email, 2=Máquina, 3=DataReg, 4=UltimaData
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
                # Verifica se a linha não está vazia e compara o nome
                if row[0] and str(row[0]).lower() == termo_busca.lower():
                    
                    # Preenche os campos com os dados do Excel
                    entry_email.insert(0, str(row[1]) if row[1] else "")
                    entry_maquina.insert(0, str(row[2]) if row[2] else "")
                    entry_data_reg.insert(0, str(row[3]) if row[3] else "")
                    entry_ultima_data.insert(0, str(row[4]) if row[4] else "")
                    
                    encontrado = True
                    break # Para de procurar ao encontrar
            
            if not encontrado:
                messagebox.showinfo("Resultado", "Colaborador não encontrado.")
                
        except Exception as e:
            messagebox.showerror("Erro Crítico", f"Erro ao ler planilha: {e}")
        
        # Trava os campos novamente (apenas leitura)
        desativar_campos()

    # --- FUNÇÕES VISUAIS ---
    def ativar_campos():
        entry_email.configure(state=tk.NORMAL)
        entry_maquina.configure(state=tk.NORMAL)
        entry_data_reg.configure(state=tk.NORMAL)
        entry_ultima_data.configure(state=tk.NORMAL)

    def desativar_campos():
        entry_email.configure(state=tk.DISABLED)
        entry_maquina.configure(state=tk.DISABLED)
        entry_data_reg.configure(state=tk.DISABLED)
        entry_ultima_data.configure(state=tk.DISABLED)

    def limpar_resultados():
        entry_email.delete(0, 'end')
        entry_maquina.delete(0, 'end')
        entry_data_reg.delete(0, 'end')
        entry_ultima_data.delete(0, 'end')

    # --- INTERFACE  ---
    frame1 = tk.LabelFrame(janela, text="Buscar Dados em Relatorio.xlsx")
    frame1.pack(expand='yes', fill='both', padx=20, pady=20)

    # Rótulos (Labels)
    tk.Label(frame1, text="Nome Colaborador: ").place(x=50, y=60)
    tk.Label(frame1, text="E-mail: ").place(x=50, y=100)
    tk.Label(frame1, text="Máquina: ").place(x=50, y=140)
    tk.Label(frame1, text="Data Registro: ").place(x=50, y=180)
    tk.Label(frame1, text="Última Data: ").place(x=50, y=220)

    # Campos de Entrada
    entry_colaborador = tk.Entry(frame1, width=40)
    entry_colaborador.place(x=200, y=60)

    entry_email = tk.Entry(frame1, width=40)
    entry_email.place(x=200, y=100)

    entry_maquina = tk.Entry(frame1, width=40)
    entry_maquina.place(x=200, y=140)

    entry_data_reg = tk.Entry(frame1, width=40)
    entry_data_reg.place(x=200, y=180)
    
    entry_ultima_data = tk.Entry(frame1, width=40)
    entry_ultima_data.place(x=200, y=220)

    # Estado inicial
    desativar_campos()

    # Botão
    tk.Button(frame1, text="Pesquisar", command=submit, height=2, width=15).place(x=500, y=60)

    if not master_window:
        janela.mainloop()

# Teste rápido
if __name__ == "__main__":
    abrir_janela_busca()