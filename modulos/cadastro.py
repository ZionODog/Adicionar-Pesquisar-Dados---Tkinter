import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import openpyxl
import os

def abrir_janela_cadastro(master_window=None):
    if master_window:
        root = tk.Toplevel(master_window)
    else:
        root = tk.Tk()
    
    root.title("Adicionar Dados")
    
    # --- CAMINHOS ---
    diretorio_base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    caminho_tema_dark = os.path.join(diretorio_base, 'theme', 'forest-dark.tcl')
    caminho_tema_light = os.path.join(diretorio_base, 'theme', 'forest-light.tcl')
    caminho_excel = os.path.join(diretorio_base, 'data', 'Relatorio.xlsx')

    # --- VERIFICAÇÃO ---
    # Se o arquivo não existir, cria ele com os cabeçalhos certos
    if not os.path.exists(caminho_excel):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ativos" # Nome da aba
        # Cria os cabeçalhos se o arquivo for novo
        ws.append(["Colaborador", "E-mail", "Máquina", "Data de Registro", "Última Data"])
        wb.save(caminho_excel)

    # --- ESTILO ---
    style = ttk.Style(root)
    try:
        # Tenta carregar os temas se os arquivos existirem
        if os.path.exists(caminho_tema_light) and os.path.exists(caminho_tema_dark):
            root.tk.call("source", caminho_tema_light)
            root.tk.call("source", caminho_tema_dark)
            style.theme_use("forest-dark")
    except:
        pass # Se der erro no tema, abre com o visual padrão do Windows

    def toggle_mode():
        if mode_switch.instate(["selected"]):
            style.theme_use("forest-light")
        else:
            style.theme_use("forest-dark")

    def load_data():
        if not os.path.exists(caminho_excel):
            return

        workbook = openpyxl.load_workbook(caminho_excel)
        # Tenta pegar a aba "Ativos", senão pega a primeira que tiver
        if "Ativos" in workbook.sheetnames:
            sheet = workbook["Ativos"]
        else:
            sheet = workbook.active

        list_values = list(sheet.values)
        
        # Limpa a tabela antes de recarregar
        for item in treeview.get_children():
            treeview.delete(item)

        if list_values:
            # Configura cabeçalhos baseados na primeira linha do Excel
            for col_name in list_values[0]:
                treeview.heading(col_name, text=col_name)

            # Insere os dados
            for value_tuple in list_values[1:]:
                treeview.insert('', tk.END, values=value_tuple)
            
    def insert_row():
        colaborador = name_entry.get()
        email = email_entry.get()
        maquina = maquina_entry.get()
        data_registro = datar_entry.get()
        ultima_data = ultimadata_entry.get()

        if not colaborador or colaborador == "Colaborador":
            messagebox.showwarning("Aviso", "Nome do colaborador é obrigatório!")
            return

        try:
            workbook = openpyxl.load_workbook(caminho_excel)
            if "Ativos" in workbook.sheetnames:
                sheet = workbook["Ativos"]
            else:
                sheet = workbook.active
                
            row_values = [colaborador, email, maquina, data_registro, ultima_data]
            sheet.append(row_values)
            workbook.save(caminho_excel)

            # Atualiza a visualização na hora
            treeview.insert('', tk.END, values=row_values)
            
            limpar_campos()
        except PermissionError:
            messagebox.showerror("Erro", "O arquivo Excel está aberto! Feche-o para salvar novos dados.")

    def limpar_campos():
        # Reseta os campos para os placeholders
        name_entry.delete(0, "end"); name_entry.insert(0, "Colaborador")
        email_entry.delete(0, "end"); email_entry.insert(0, "E-mail")
        maquina_entry.delete(0, "end"); maquina_entry.insert(0, "Máquina")
        datar_entry.delete(0, "end"); datar_entry.insert(0, "Data de Registro")
        ultimadata_entry.delete(0, "end"); ultimadata_entry.insert(0, "Última Data")

    # --- INTERFACE ---
    frame = ttk.Frame(root)
    frame.pack()
    
    widgets_frame = ttk.LabelFrame(frame, text="Novo Cadastro")
    widgets_frame.grid(row=0, column=0, padx=20, pady=20)

    name_entry = ttk.Entry(widgets_frame); name_entry.grid(row=0, column=0, sticky="ew", pady=5)
    name_entry.insert(0, "Colaborador")
    # Limpa o texto ao clicar (placeholder)
    name_entry.bind("<FocusIn>", lambda e: name_entry.delete('0', 'end') if name_entry.get() == "Colaborador" else None)
    
    email_entry = ttk.Entry(widgets_frame); email_entry.grid(row=1, column=0, sticky="ew", pady=5)
    email_entry.insert(0, "E-mail")
    email_entry.bind("<FocusIn>", lambda e: email_entry.delete('0', 'end') if email_entry.get() == "E-mail" else None)

    maquina_entry = ttk.Entry(widgets_frame); maquina_entry.grid(row=2, column=0, sticky="ew", pady=5)
    maquina_entry.insert(0, "Máquina")
    maquina_entry.bind("<FocusIn>", lambda e: maquina_entry.delete('0', 'end') if maquina_entry.get() == "Máquina" else None)

    datar_entry = ttk.Entry(widgets_frame); datar_entry.grid(row=3, column=0, sticky="ew", pady=5)
    datar_entry.insert(0, "Data de Registro")
    datar_entry.bind("<FocusIn>", lambda e: datar_entry.delete('0', 'end') if datar_entry.get() == "Data de Registro" else None)

    ultimadata_entry = ttk.Entry(widgets_frame); ultimadata_entry.grid(row=4, column=0, sticky="ew", pady=5)
    ultimadata_entry.insert(0, "Última Data")
    ultimadata_entry.bind("<FocusIn>", lambda e: ultimadata_entry.delete('0', 'end') if ultimadata_entry.get() == "Última Data" else None)

    button = ttk.Button(widgets_frame, text="Adicionar", command=insert_row)
    button.grid(row=5, column=0, sticky="nsew", pady=5)

    mode_switch = ttk.Checkbutton(widgets_frame, text="Tema Escuro", style="Switch", command=toggle_mode)
    mode_switch.grid(row=7, column=0, sticky="nsew", pady=10)

    # Tabela à direita
    treeFrame = ttk.Frame(frame)
    treeFrame.grid(row=0, column=1, pady=10)
    treeScroll = ttk.Scrollbar(treeFrame)
    treeScroll.pack(side="right", fill="y")
    
    cols = ("Colaborador", "E-mail", "Máquina", "Data de Registro", "Última Data")
    treeview = ttk.Treeview(treeFrame, show="headings", yscrollcommand=treeScroll.set, columns=cols, height=13)
    
    for col in cols:
        treeview.column(col, width=120)
        treeview.heading(col, text=col)
    
    treeview.pack()
    treeScroll.config(command=treeview.yview)
    
    load_data()

    if not master_window:
        root.mainloop()

if __name__ == "__main__":
    abrir_janela_cadastro()